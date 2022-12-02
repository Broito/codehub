import openpyxl
import os
import pandas as pd

# 切换工作文件夹
os.chdir(r'E:\workspace\Research_2022_CDC\excel转word') 
file = openpyxl.load_workbook('匹配及打印.xlsx')
wb = file['打印版']
df = pd.read_excel('原表(1).xlsx','【二维表】记录数据,共计(3条)')
df = df.fillna('')

def input_items(order):
    if order == 1:
        j = "B"
        i = 3
    elif order == 2:
        j = "G"
        i = 3
    elif order == 3:
        j = "B"
        i = 22
    elif order == 0:
        j = "G"
        i = 22
    wb[j+str(i)] = recordNumber
    wb[j+str(i+1)] = eventDate
    wb[j+str(i+2)] = recordedBy
    wb[j+str(i+3)] = address
    wb[j+str(i+4)] = coordinate
    wb[j+str(i+5)] = habitat
    wb[openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(j)+2)+str(i+5)] = elevation
    wb[j+str(i+6)] = habit
    wb[openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(j)+2)+str(i+6)] = heightInMeters
    wb[j+str(i+7)] = frequency
    wb[j+str(i+8)] = desc
    wb[j+str(i+9)] = individualCount
    wb[j+str(i+10)] = occurrenceRemarks

for row in df.iterrows():

    index = row[0]
    item = row[1]

    recordNumber = str(item['recordNumber'])
    eventDate = str(item['eventDate'])
    recordedBy = str(item['recordedBy'])
    address = str(item['country'])+str(item['stateProvince'])+str(item['city'])+str(item['county'])+str(item['locality'])
    coordinate = str(item['decimalLongitude'])+','+str(item['decimalLatitude'])
    habitat = str(item['habitat'])
    elevation = str(item['minimumElevationInMeters'])
    habit = str(item['habit'])
    heightInMeters = str(item['体高'])
    frequency = str(item['频度'])
    desc = str(item['果实'])+'；'+str(item['种子'])+'；'+str(item['花'])
    # 这个原表里没有对应的，暂时使用occurrenceRemarks来代替，可直接更改
    individualCount = str(item['occurrenceRemarks'])
    occurrenceRemarks = str(item['vernacularName'])+'；'+str(item['scientificName'])+'；'+str(item['recordNumber'])

    input_items((index+1)%4)
    if (index+1)%4 == 0 or (index+1) == df.shape[0]:
        output_name = f"{int((index+1)/4)}.xlsx"
        file.save('./results/'+output_name)
























